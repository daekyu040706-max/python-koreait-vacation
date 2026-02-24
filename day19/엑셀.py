# 파이썬에서 엑셀다루기 -> pip install openpyxl
from openpyxl import Workbook

# .xlsx 파일 생성(워크북생성)
wb = Workbook()

# 기본시트(sheet1)를 현재시트로 지정
ws = wb.active
ws.title = "첫번째시트" # 시트명 변경

# 엑셀의 cell에 데이터 입력하기
# 1. 좌표로 데이터 입력
ws["A1"] = "이름"
ws["B1"] = "나이"
ws["C1"] = "직업"

# 2. row와 column으로 입력
ws.cell(row=2, column=1, value="홍길동")
ws.cell(row=2, column=2, value=25)
ws.cell(row=2, column=3, value="개발자")

# 3. 리스트로 입력하기
# 마지막 row에 입력 됨
ws.append(["김길동", 30, "디자이너"])

people = [
    ["박길동", 27, "교사"],
    ["최길동", 28, "의사"],
    ["이길동", 29, "변호사"]
    ]
for p in people:
    ws.append(p)


wb.save("./sample.xlsx")